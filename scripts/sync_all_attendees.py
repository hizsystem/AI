#!/usr/bin/env python3
"""3개 채널(온오프믹스, 이벤터스, 웹사이트) 참석자를 구글시트 '참석자 명단'에 통합

사용법:
  python scripts/sync_all_attendees.py \
    --onoffmix /path/to/onoffmix.xls \
    --eventers /path/to/eventers.xlsx \
    --website /path/to/website.csv
"""

import argparse
import sys
import time

sys.path.insert(0, ".")
from scripts.sheets_auth import get_client

import csv
import xlrd

# ── 설정 ──
SHEET_URL = "https://docs.google.com/spreadsheets/d/1YtOd5pW9ECd_kQggWpk8VojUlpuMXn9PS7R1kBG4NiM/edit"
TAB_NAME = "참석자 명단"

# ── 파서: 온오프믹스 ──
def read_onoffmix(filepath):
    """온오프믹스 XLS에서 참가자 추출"""
    wb = xlrd.open_workbook(filepath, ignore_workbook_corruption=True)
    ws = wb.sheet_by_index(0)
    participants = []
    for r in range(5, ws.nrows):
        row = [ws.cell_value(r, c) for c in range(ws.ncols)]
        if not row[0] or not isinstance(row[0], (int, float)) or not row[2]:
            continue
        participants.append({
            "name": str(row[2]).strip(),
            "org": str(row[3]).strip() if row[3] else "",
            "phone": str(row[4]).strip(),
            "email": str(row[5]).strip(),
            "count": int(row[6]) if isinstance(row[6], (int, float)) else 1,
            "confirmed": "Y" if row[8] == "Y" else "N",
            "date": str(row[10]).strip(),
            "channel": "온오프믹스",
        })
    return participants


# ── 파서: 이벤터스 ──
def read_eventers(filepath):
    """이벤터스 XLSX에서 참가자 추출"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    participants = []
    for r in range(5, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        if not seq:
            continue
        name = ws.cell(r, 4).value or ""
        email = ws.cell(r, 5).value or ""
        phone_raw = ws.cell(r, 7).value or ""
        org = ws.cell(r, 8).value or ""
        position = ws.cell(r, 9).value or ""
        status = ws.cell(r, 10).value or ""
        count = ws.cell(r, 13).value or 1
        date_str = ws.cell(r, 3).value or ""

        # 전화번호 정리 (이벤터스는 '-' 없이 올 수 있음)
        phone = str(phone_raw).strip()
        if phone and not phone.startswith("0"):
            phone = "0" + phone

        # 날짜 정리 (2026-03-11 14:23:18 형태)
        date_only = str(date_str).split(" ")[0] if date_str else ""

        participants.append({
            "name": str(name).strip(),
            "org": str(org).strip(),
            "phone": phone,
            "email": str(email).strip(),
            "count": int(count) if count else 1,
            "confirmed": "Y" if "확정" in str(status) else "N",
            "date": date_only,
            "channel": "이벤터스",
            "position": str(position).strip(),
        })
    return participants


# ── 파서: 웹사이트 CSV ──
def read_website_csv(filepath):
    """웹사이트(Supabase) 내보내기 CSV에서 참가자 추출.

    CSV 컬럼: 기업명,참석자명,직책,이메일,연락처,창업단계,신청상태,결제상태,신청일시

    중복/실패 건 처리:
    - 같은 사람이 여러 번 신청한 경우 confirmed+done 우선, 그 다음 최신 건
    - 결제 실패(failed)만 있고 성공 건이 없는 경우도 포함 (pending 처리)
    """
    with open(filepath, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        raw = list(reader)

    # 이메일 기준으로 그룹핑하여 최적 건 선택
    by_email = {}
    for row in raw:
        email = (row.get("이메일") or "").strip().lower()
        if not email:
            continue
        if email not in by_email:
            by_email[email] = []
        by_email[email].append(row)

    participants = []
    for email, rows in by_email.items():
        # 우선순위: confirmed+done > confirmed+* > pending+done > 나머지 (최신 우선)
        def score(r):
            s = 0
            if r.get("신청상태") == "confirmed":
                s += 10
            if r.get("결제상태") == "done":
                s += 5
            elif r.get("결제상태") == "ready":
                s += 2
            return s

        best = max(rows, key=score)

        # 날짜 정리: "2026. 3. 22. PM 7:32:17" → "2026-03-22"
        date_raw = best.get("신청일시", "")
        date_only = ""
        if date_raw:
            parts = date_raw.replace(".", "").split()
            if len(parts) >= 3:
                try:
                    y, m, d = parts[0], parts[1], parts[2]
                    date_only = f"{y}-{int(m):02d}-{int(d):02d}"
                except (ValueError, IndexError):
                    date_only = date_raw

        confirmed = "Y" if best.get("신청상태") == "confirmed" else "N"

        participants.append({
            "name": (best.get("참석자명") or "").strip(),
            "org": (best.get("기업명") or "").strip(),
            "phone": (best.get("연락처") or "").strip(),
            "email": (best.get("이메일") or "").strip(),
            "count": 1,
            "confirmed": confirmed,
            "date": date_only,
            "channel": "웹사이트",
            "position": (best.get("직책") or "").strip(),
            "payment": best.get("결제상태", ""),
        })
    return participants


# ── 중복 제거 ──
def deduplicate(all_participants):
    """이메일 또는 전화번호 기준 중복 제거 (첫 등장 채널 우선)"""
    seen_emails = {}
    seen_phones = {}
    unique = []
    dupes = []

    for p in all_participants:
        email = p["email"].lower().strip()
        phone = p["phone"].replace("-", "").strip()

        dup_of = None
        if email and email in seen_emails:
            dup_of = seen_emails[email]
        elif phone and phone in seen_phones:
            dup_of = seen_phones[phone]

        if dup_of:
            dupes.append((p, dup_of))
            continue

        if email:
            seen_emails[email] = p
        if phone:
            seen_phones[phone] = p
        unique.append(p)

    return unique, dupes


# ── 구글시트 업데이트 ──
def update_sheet(participants, dupes):
    """구글시트 '참석자 명단' 탭을 전체 갱신"""
    client = get_client()
    sh = client.open_by_url(SHEET_URL)

    # 탭 가져오기 또는 생성
    try:
        ws = sh.worksheet(TAB_NAME)
    except Exception:
        ws = sh.add_worksheet(title=TAB_NAME, rows=500, cols=12)

    # 헤더 설정
    headers = [
        "No", "채널", "이름", "소속", "직책",
        "전화", "이메일", "신청인원", "참여확정",
        "신청일", "비고"
    ]

    # 데이터 행 구성
    rows = [headers]
    for i, p in enumerate(participants, 1):
        rows.append([
            i,
            p["channel"],
            p["name"],
            p["org"],
            p.get("position", ""),
            p["phone"],
            p["email"],
            p["count"],
            p["confirmed"],
            p["date"],
            "",
        ])

    # 기존 데이터 클리어 후 전체 쓰기
    ws.clear()
    time.sleep(1)

    ws.update(values=rows, range_name="A1", value_input_option="USER_ENTERED")
    time.sleep(1)

    # 헤더 서식
    ws.batch_format([
        {"range": "A1:K1", "format": {
            "textFormat": {
                "bold": True,
                "foregroundColorStyle": {"rgbColor": {"red": 1, "green": 1, "blue": 1}},
            },
            "backgroundColor": {"red": 0.2, "green": 0.3, "blue": 0.6},
            "horizontalAlignment": "CENTER",
        }},
    ])

    # 채널별 색상 (선택)
    channel_colors = {
        "온오프믹스": {"red": 0.95, "green": 0.95, "blue": 1.0},
        "이벤터스": {"red": 0.95, "green": 1.0, "blue": 0.95},
        "웹사이트": {"red": 1.0, "green": 0.95, "blue": 0.95},
    }
    formats = []
    for i, p in enumerate(participants, 2):  # 데이터는 2행부터
        color = channel_colors.get(p["channel"])
        if color:
            formats.append({
                "range": f"A{i}:K{i}",
                "format": {"backgroundColor": color},
            })
    if formats:
        # gspread batch_format은 한번에 최대 100개까지
        for chunk_start in range(0, len(formats), 100):
            ws.batch_format(formats[chunk_start:chunk_start + 100])
            time.sleep(1)

    return ws


def main():
    parser = argparse.ArgumentParser(description="3채널 참석자 통합 → 구글시트")
    parser.add_argument("--onoffmix", help="온오프믹스 XLS 파일 경로")
    parser.add_argument("--eventers", help="이벤터스 XLSX 파일 경로")
    parser.add_argument("--website", help="웹사이트 신청자 CSV 파일 경로")
    parser.add_argument("--dry-run", action="store_true", help="시트 업데이트 없이 확인만")
    args = parser.parse_args()

    all_participants = []

    # 1. 온오프믹스
    if args.onoffmix:
        onoffmix = read_onoffmix(args.onoffmix)
        print(f"1. 온오프믹스: {len(onoffmix)}명")
        all_participants.extend(onoffmix)
    else:
        print("1. 온오프믹스: (파일 미지정, 건너뜀)")

    # 2. 이벤터스
    if args.eventers:
        eventers = read_eventers(args.eventers)
        print(f"2. 이벤터스: {len(eventers)}명")
        all_participants.extend(eventers)
    else:
        print("2. 이벤터스: (파일 미지정, 건너뜀)")

    # 3. 웹사이트 CSV
    if args.website:
        website = read_website_csv(args.website)
        print(f"3. 웹사이트: {len(website)}명")
        all_participants.extend(website)
    else:
        print("3. 웹사이트: (파일 미지정, 건너뜀)")

    print(f"\n--- 총 수집: {len(all_participants)}명 ---")

    # 4. 중복 제거
    unique, dupes = deduplicate(all_participants)
    print(f"중복 제거 후: {unique}명" if isinstance(unique, int) else f"중복 제거 후: {len(unique)}명")

    if dupes:
        print(f"\n중복 {len(dupes)}건:")
        for dup, original in dupes:
            print(f"  - {dup['name']} ({dup['channel']}) ← 이미 {original['channel']}에서 등록")

    # 채널별 집계
    channel_counts = {}
    for p in unique:
        ch = p["channel"]
        channel_counts[ch] = channel_counts.get(ch, 0) + 1
    print(f"\n채널별 (중복 제거 후):")
    for ch, cnt in channel_counts.items():
        print(f"  {ch}: {cnt}명")
    print(f"  합계: {len(unique)}명")

    if args.dry_run:
        print("\n(dry-run: 시트 업데이트 건너뜀)")
        return

    # 5. 구글시트 업데이트
    print(f"\n구글시트 '{TAB_NAME}' 업데이트 중...")
    ws = update_sheet(unique, dupes)
    print(f"\n✅ 완료! {len(unique)}명 → 구글시트 '{TAB_NAME}'")
    print(f"📊 시트: {SHEET_URL}")


if __name__ == "__main__":
    main()
